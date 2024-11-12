"""
**********************************************************************************

TestIDManager.py
Created by: Debashis Das
Argonne National Laboratory
Transportation and Power Systems Division

**********************************************************************************
  
Description:
------------
The methods available from this class are the following:
- get_cycles_list(): Method to get list of cycles from config file
- is_description_row(row): Method dentify description rows
- dict_to_df(dictionary): Method to create pandas dataframe from dictionary
********
"""

import pandas as pd
import json
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from TdmsFileManager import TdmsFileManager
from openpyxl.styles import Border, Side, Font, Alignment

class TestIDManager:
    def __init__(self, vehicle_name, drive_cycle, platform, tdms_data_directory):
        
        configFile = open(self.get_config_file(vehicle_name), 'r', encoding='utf-8')
        self.config = (json.load(configFile))
        configFile.close()
        self.platform = platform
        self.tdms_data_directory = tdms_data_directory
        print(self.output_file_path)
        
        self.get_cycles_list(drive_cycle)
        self.output_sheet_name = "72F_Cycle_Data_Hioki"
        self.iteration_keys = ["1st","2nd","3rd","4th","5th","6th","7th","8th","9th","10th"]
        self.desired_test_id_list = []
        self.depletion_test_id_list = []
        self.test_id_list_category_1st, self.test_id_list_category_2nd, self.test_id_list_category_3rd, self.test_id_list_category_4th, self.test_id_list_category_5th = ([] for i in range(5))

    def get_config_file(self, vehicle_name):
        if vehicle_name == "2020 Tesla Model 3":
            config_file_name = os.path.join("config-files", "configuration_tesla.json")
            self.output_file_path = os.path.join("Data", "Tesla-Model3", "2020-tesla-model3-uncertainity-analysis.xlsx")

        elif vehicle_name == "2020 Chevrolet Bolt":
            config_file_name = os.path.join("config-files", "configuration_bolt.json")
            self.output_file_path = os.path.join("Data", "Chevrolet_Bolt", "2020-chevy-bolt-uncertainity-analysis.xlsx")

        elif vehicle_name == "2019 Nissan Leaf":
            config_file_name = os.path.join("config-files", "configuration_leaf.json")
            self.output_file_path = os.path.join("Data", "Nissan-Leaf", "2019-nissan-leaf-uncertainity-analysis.xlsx")
            
        else:
            raise ValueError("Unknown vehicle name")

        return config_file_name


    def get_cycles_list(self, drive_cycle):
        """
        Method to get list of cycles from config file
        """
        if drive_cycle == "MCT":
            self.cycles_list = ['UDDS 1 , Combined', 'UDDS 2, combined', 'UDDS 3, combined', 'UDDS 4, combined', 'HWY 1', 'HWY', 'US06 1, combined', 'US06 2, combined', 'SSS 65mph depletion', 'Charge L2 23C']
            self.sub_cycle_names_list = ['UDDS 1 , 505', 'UDDS 1 ', 'UDDS 1 , Combined', 'HWY 1', 'UDDS 2, 505', 'UDDS 2', 'UDDS 2, combined', 'US06 1, city', 'US06 1, hwy', 'US06 1, combined', 'SSS 65mph depletion', 'US06 2, city', 'US06 2, hwy', 'US06 2, combined', 'UDDS 3, 505', 'UDDS 3', 'UDDS 3, combined', 'HWY', 'UDDS 4, 505', 'UDDS 4', 'UDDS 4, combined', 'SSS 65mph depletion', 'Charge L2 23C']
            self.cycles_list_dic = [{'UDDS': ['UDDS 1 , Combined', 'UDDS 2, combined', 'UDDS 3, combined', 'UDDS 4, combined']}, {
                'HWY': ['HWY 1', 'HWY']}, {'US06': ['US06 1, combined', 'US06 2, combined']}, {'SSS 65mph': ['SSS 65mph depletion']}, {'Charge L2 23C': ['Charge L2 23C']}]
            
    def check_description_row(self, row):
        """
        Define a function to identify description rows
        """
        return pd.isna(row['Test Time']) and pd.isna(row['Date'])
    
    def matching_cycles(self, dataframe):
        """
        Method to find index number in the test file dataframe that matches desired drive cycle (e.g., MCT)
        """
        
        # create dictionaries dynamically for cycle types (e.g., UDDS_dictionary, US06_dictionary, etc.)
        cycle_type_dicts = {}

        for cycle_type in self.cycles_list_dic:
            for cycle_type_key, cycle_type_value in cycle_type.items():
                dict_name = f"{cycle_type_key}_dictionary" 
                cycle_type_dicts[dict_name] = {}                

        max_sub_phase_number = self.config["MaxSubPhase"]
        sub_phase_number = 1
        sub_cycle_name_counter = 0
        iteration_index = 0
        desired_index_list  =[]
        sub_cycle_finding_status = True
            
        for index, row in dataframe.loc[:].iterrows():
        
            if row['Cycle'] == self.sub_cycle_names_list[sub_cycle_name_counter]:
                sub_cycle_name_counter += 1
                sub_cycle_finding_status = True            
                
                if row['Cycle'] in self.cycles_list:
                    desired_index_list.append(index)            
                    
            else: sub_cycle_finding_status = False
                        
            if sub_cycle_finding_status and sub_cycle_name_counter == len(self.sub_cycle_names_list):
                sub_cycle_finding_status = False

                # iterating both cycle_type dictionary and cycle types simultaneouly
                for (dict_name, dictionary), cycle_type in zip(cycle_type_dicts.items(), self.cycles_list_dic):
                    for key, values in cycle_type.items():
                        for sub_cycle_name in values:
                            if sub_phase_number > max_sub_phase_number:
                                sub_phase_number = 1
                            
                            if iteration_index < len(self.iteration_keys):
                                self.populate_dictionary(dataframe, desired_index_list, sub_cycle_name, sub_phase_number, self.iteration_keys[iteration_index], dictionary, dict_name)
                            
                            else:
                                print("Not enough iteration keys for all cycle types")
                                break
                            
                            sub_phase_number += 1
                            
                iteration_index += 1
                
            if not(sub_cycle_finding_status)  and sub_cycle_name_counter > 1:
                sub_cycle_name_counter = 0
                desired_index_list.clear()
        
        return cycle_type_dicts
    
    def populate_dictionary(self,df, index_list, desired_subcycle_name, sub_phase_number, iteration_key, dictionary, dictionary_name):
        """
        Method to create lists dynamically and append requied data into those lists
        """
        # create lists dynamically
        lists_dict = {f"{field}": [] for field in self.config['DataFields']}
        
        for index in index_list:
            if df.loc[index, 'Cycle'] == desired_subcycle_name:
                for field in self.config['DataFields']:
                    if field in df.columns:
                        lists_dict[f"{field}"].append(df.loc[index, field])

                    # else:
                    #     print(f"Field '{field}' does not exist in the dataframe. Skipping this field.")

                    if field == "Test ID [#]" and desired_subcycle_name == "SSS 65mph depletion":
                        self.depletion_test_id_list.append(df.loc[index, field])

                    if field == "Test ID [#]" and desired_subcycle_name == "Charge L2 23C":
                        self.depletion_test_id_list.append(df.loc[index, field])

                    elif field == "Test ID [#]":
                        self.desired_test_id_list.append(df.loc[index, field])            

                if desired_subcycle_name in ['UDDS 1 , Combined', 'UDDS 2, combined', 'UDDS 3, combined', 'UDDS 4, combined']:
                    lists_dict["Sub-Phase"].append(sub_phase_number)

                if iteration_key == '1st':
                    self.test_id_list_category_1st.append(df.loc[index, "Test ID [#]"])

                elif iteration_key == '2nd':
                    self.test_id_list_category_2nd.append(df.loc[index, "Test ID [#]"])

                elif iteration_key == '3rd':
                    self.test_id_list_category_3rd.append(df.loc[index, "Test ID [#]"])

                elif iteration_key == '4th':
                    self.test_id_list_category_4th.append(df.loc[index, "Test ID [#]"])

                elif iteration_key == '5th':
                    self.test_id_list_category_5th.append(df.loc[index, "Test ID [#]"])

        self.update_dict(iteration_key, lists_dict, dictionary, dictionary_name)   
    
    def dict_to_df(self, dictionary):
        """
        Method to create pandas dataframe from dictionary 
        """
        data_list = []
        for key, entries in dictionary.items():
            for entry in entries:
                # Find the maximum length of lists in the entry
                max_length = max(len(v) for v in entry.values())
                for i in range(max_length):
                    data_entry = {}
                    for k, v in entry.items():
                        data_entry[k] = v[i] if i < len(v) else None
                    data_entry['Category'] = key
                    data_list.append(data_entry)

        df = pd.DataFrame(data_list)
        # Reorder columns to place 'Category' at the start
        columns = ['Category'] + [col for col in df.columns if col != 'Category']
        df = df[columns]
        return df  
        
    def merge_category_cells(self, df, start_row, sheet):
        """
        Function to merge cells in the 'Category' column
        """
        current_category = None
        start_merge = None
        
        for idx, row in df.iterrows():
            category = row['Category']
            row_num = start_row + idx
            
            if category == current_category:
                continue
            
            if current_category is not None:
                end_merge = row_num
                sheet.merge_cells(start_row=start_merge, start_column=1, end_row=end_merge, end_column=1)
            current_category = category
            start_merge = row_num + 1
            
        # Merge the last group
        if start_merge is not None:
            sheet.merge_cells(start_row=start_merge, start_column=1, end_row=row_num + 1, end_column=1)
            
    def update_dict(self, key, value, dictionary, dictionary_name):
        """
        Method to update msg_count_dictionary
        """
        if key in dictionary:
            dictionary[key].append(value)
        else:
            dictionary[key] = [value]
            
        # print(f"Data in {dictionary_name} is:\n{dictionary}")        
    def write_in_excel_file(self, filtered_dataframe):
        """
        """
        output_dir = os.path.dirname(self.output_file_path)
        if not os.path.isdir(output_dir):
            os.makedirs(output_dir)
            print(f"The directory '{output_dir}' was created.")
        else:
            print(f"The directory '{output_dir}' already exists.")
            
        cycle_type_dicts = self.matching_cycles(filtered_dataframe)
        
        # Create a Pandas Excel writer using openpyxl as the engine
        with pd.ExcelWriter(self.output_file_path, engine='openpyxl') as writer:
            start_row = 0
            for dict_name, dictionary in cycle_type_dicts.items():
                # Convert dictionary to DataFrame
                df = self.dict_to_df(dictionary)
                # Round the DataFrame to 2 decimal places
                df = df.round(2)
                
                # Add dictionary name before DataFrame
                worksheet = writer.book.create_sheet(self.output_sheet_name) if start_row == 0 else writer.sheets[self.output_sheet_name]
                worksheet.append([dict_name.split('_')[0].upper()]) if start_row == 0 else worksheet.cell(row=start_row + 1, column=1, value=dict_name.split('_')[0].upper())
                
                # Write the DataFrame to the sheet starting at the next row
                df.to_excel(writer, sheet_name=self.output_sheet_name, startrow=start_row + 1, index=False)
                # Apply styling for each DataFrame written
                workbook = writer.book
                self.style_dataframe(workbook, self.output_sheet_name, start_row=start_row + 2, dataframe=df)
                # Update start_row for the next dictionary
                start_row += len(df) + 3

        # Load the workbook and access the sheet
        workbook = load_workbook(self.output_file_path)
        sheet = workbook[self.output_sheet_name]
        
        # Merge cells in 'Category' column for each DataFrame
        start_row = 2
        for dict_name, dictionary in cycle_type_dicts.items():
            df = self.dict_to_df(dictionary)
            self.merge_category_cells(df, start_row, sheet)
            start_row += len(df) + 3

        # Save the updated workbook
        workbook.save(self.output_file_path)
        print(f"File created successfully: {self.output_file_path}")


    def style_dataframe(self, wb, sheet_name, start_row, dataframe):
        # Load the specified sheet
        sheet = wb[sheet_name]

        # Define border style
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Bold font for column headers
        bold_font = Font(bold=True)

        # Set text alignment to wrap text
        wrap_alignment = Alignment(wrap_text=True)

        # Get the number of rows and columns in the DataFrame
        n_rows, n_cols = dataframe.shape
        
        # Apply styles to the column headers (first row of the DataFrame or summary)
        for col_idx in range(1, n_cols + 1):
            cell = sheet.cell(row=start_row, column=col_idx)
            cell.font = bold_font  # Bold the column header
            cell.border = thin_border  # Apply border to the column header
            cell.alignment = wrap_alignment
        
        # Apply borders only to the range where the DataFrame has actual data
        for row_idx in range(start_row + 1, start_row + n_rows + 1):
            for col_idx in range(1, n_cols + 1):
                # Get the actual value from the DataFrame for this cell
                # cell_value = dataframe.iloc[row_idx - start_row - 1, col_idx - 1]
                
                # # Only apply border if the cell has non-NaN, non-empty data
                # if pd.notna(cell_value) and cell_value != "":
                cell = sheet.cell(row=row_idx, column=col_idx)
                cell.border = thin_border  # Apply border to each non-empty cell
                cell.alignment = wrap_alignment

                     
    def manage_test_data(self):
        """
        """
        
        data_frame = pd.read_excel(self.config["InputFileName"], sheet_name = self.config['InputSheetName'],  skiprows = self.config['NoOfSkipRows'])        
        filtered_dataframe = data_frame[~data_frame.apply(self.check_description_row, axis=1)] # Apply the filter to exclude description rows
        
        self.write_in_excel_file(filtered_dataframe)
        self.test_id_list_category_1st = sorted(list(set(self.test_id_list_category_1st)))
        self.test_id_list_category_2nd = sorted(list(set(self.test_id_list_category_2nd)))
        self.test_id_list_category_3rd = sorted(list(set(self.test_id_list_category_3rd)))
        self.test_id_list_category_4th = sorted(list(set(self.test_id_list_category_4th)))
        self.test_id_list_category_5th = sorted(list(set(self.test_id_list_category_5th)))
        self.desired_test_id_list =  sorted(list(set(self.desired_test_id_list)))
        self.depletion_test_id_list = sorted(list(set(self.depletion_test_id_list)))
                
        tdms_file_manager = TdmsFileManager("TDMS File Manager object", self.platform, self.output_file_path, self.tdms_data_directory)
        tdms_file_manager.set_test_id_list(self.test_id_list_category_1st, self.test_id_list_category_2nd, self.test_id_list_category_3rd, self.test_id_list_category_4th, self.test_id_list_category_5th)
        tdms_file_manager.manage_tdms_file(self.desired_test_id_list)
        tdms_file_manager.manage_depletion_tdms_file(self.depletion_test_id_list)
        del tdms_file_manager


'''##############################################
                   Unit testing
##############################################'''
if __name__ == "__main__":
    # test_id_manager = TestIDManager("2019 Nissan Leaf", "MCT", "Linux", "AMTL-Test-Data")
    test_id_manager = TestIDManager("2020 Tesla Model 3", "MCT", "Windows", "AMTL-Test-Data")
    test_id_manager.manage_test_data()