"""
**********************************************************************************

data-processing.py
Created by: Debashis Das
Argonne National Laboratory
Transportation and Power Systems Division

**********************************************************************************
  
Description:
------------
The methods available from this class are the following:
- get_cycles_list(config): Method to get list of cycles from config file
- is_description_row(row): Method dentify description rows
- dict_to_df(dictionary): Method to create pandas dataframe from dictionary
********
"""

import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from TestIDManager import TestIDManager



    





    

def main():
    
    # Read the config file into a json object:
    configFile = open("config-files/configuration.json", 'r')
    config = (json.load(configFile))
    configFile.close()
    
    test_id_manager = TestIDManager(config)
    
    # output_file_path = config["OutputFileName"]
    # output_sheet_name = config['OutputSheetName']
    
    # data_frame = pd.read_excel(config["InputFileName"], sheet_name = config['InputSheetName'],  skiprows = config['NoOfSkipRows'])

    # Apply the filter to exclude description rows
    # filtered_df = data_frame[~data_frame.apply(is_description_row, axis=1)]
    
    # # Reading column names
    # column_names = filtered_df.columns

    # # Display column names
    # print("Column names:", column_names)
    
    # cycles_list = get_cycles_list(config)
    # cycle_type_dicts = matching_cycles(config, filtered_df, cycles_list) 
 
    # for dictionary_name, dictionary in cycle_type_dicts.items():
    #     print(f"Data in {dictionary_name} is:\n{dictionary}\n\n")
       
    # # Create a Pandas Excel writer using openpyxl as the engine
    # with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    #     start_row = 0
    #     for dict_name, dictionary in cycle_type_dicts.items():
    #         # Convert dictionary to DataFrame
    #         df = dict_to_df(dictionary)
            
    #         # Add dictionary name before DataFrame
    #         worksheet = writer.book.create_sheet(output_sheet_name) if start_row == 0 else writer.sheets[output_sheet_name]
    #         worksheet.append([dict_name.split('_')[0].upper()]) if start_row == 0 else worksheet.cell(row=start_row + 1, column=1, value=dict_name.split('_')[0].upper())
            
    #         # Write the DataFrame to the sheet starting at the next row
    #         df.to_excel(writer, sheet_name=output_sheet_name, startrow=start_row + 1, index=False)
            
    #         # Update start_row for the next dictionary
    #         start_row += len(df) + 3

    # # Load the workbook and access the sheet
    # workbook = load_workbook(output_file_path)
    # sheet = workbook[output_sheet_name]
    
    # # Merge cells in 'Category' column for each DataFrame
    # start_row = 2
    # for dict_name, dictionary in cycle_type_dicts.items():
    #     df = dict_to_df(dictionary)
    #     merge_category_cells(df, start_row, sheet)
    #     start_row += len(df) + 3

    # # Save the updated workbook
    # workbook.save(output_file_path)
    # print(f"File created successfully: {output_file_path}")
    
    test_id_manager.manage_test_id()

if __name__ == "__main__":
    main()  