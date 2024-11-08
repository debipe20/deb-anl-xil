import numpy as np
import os
import openpyxl
import pandas as pd
from nptdms import TdmsFile
from openpyxl.styles import Border, Side, Font

class TdmsFileManager:
    def __init__(self, object_name, platform, output_file_path, tdms_directory):
        self.object_name = object_name
        self.platform = platform
        self.output_file_path = output_file_path
        # Define the path to your TDMS file using a raw string literal (r"...")
        self.set_tdms_data_directory(tdms_directory)
    
    def set_tdms_data_directory(self, tdms_directory):
        base_directory = os.path.expanduser("~")  # Gets the user's home directory
        if self.platform == "Linux":
            self.tdms_data_directory = os.path.join(base_directory, tdms_directory)
        else:  # For Windows
            self.tdms_data_directory = os.path.join(base_directory, "Documents", "Data", tdms_directory)
            
    
    def manage_tdms_file(self, test_ID_list):
        test_ID_list = [62005016]
        # Loop through each test ID in the list
        for test_id in test_ID_list:
            
            self.tdms_file_path = self.tdms_data_directory + f"/{test_id} Test Data.tdms"
            
            # print("TDMS file path is:\n", self.tdms_file_path)
            tdms_file = TdmsFile.read(self.tdms_file_path, memmap_dir=None)
            
            # Get the DataFrame and summary data
            group_channel_dataframe = self.get_data_group_channel_dataframe(tdms_file)
            summary_data = self.get_summary_table(group_channel_dataframe)
            
            # First, write the summary table to the sheet
            sheet_name = "wh_cal_" + str(test_id)
            self.add_summary_table(sheet_name, summary_data)
            
            # Write the DataFrame below the summary table using pd.ExcelWriter
            with pd.ExcelWriter(self.output_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Write the DataFrame to the specified sheet, below the summary table
                group_channel_dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(summary_data) + 2)

            # Load workbook to apply formatting
            wb = openpyxl.load_workbook(self.output_file_path)
            
            # Apply borders and bold column headers to the summary table
            self.style_dataframe(wb, sheet_name, start_row=1, dataframe=pd.DataFrame(summary_data))

            # Apply borders and bold column headers after writing the DataFrame
            self.style_dataframe(wb, sheet_name, start_row=len(summary_data) + 2, dataframe=group_channel_dataframe)

            # Save and close the workbook
            wb.save(self.output_file_path)
            wb.close()
            
            print(f"Data successfully written to sheet '{sheet_name}' in {self.output_file_path}")


    def fill_cumulative_list(self, source_list):
        cumulative_list = []
        for i in range(len(source_list)):
            if i == 0:
                cumulative_list.append((source_list[i] * 0.1) / 3600)
            else:
                cumulative_list.append(cumulative_list[i-1] + (source_list[i] * 0.1) / 3600)
        return cumulative_list

    def get_data_group_channel_dataframe(self, tdms_file):
            
        group_channel_dataframe = pd.DataFrame()
        no_cycle_wh, udds1_wh, udds2_wh, highway_wh, us06_wh = ([] for i in range(5))
        # Access the 'Data' group
        group_data = tdms_file["Data"]

        # Read the DAQ_Time[s] and P2 channels
        daq_time = group_data["DAQ_Time[s]"].data
        p2_data = group_data["P2"].data
        exhaust_bag = group_data["Exhaust_Bag"].data

        no_cycle = [p2_data[i] if exhaust_bag[i] == 0 else 0 for i in range(len(p2_data))]
        udds1_w =  [p2_data[i] if (exhaust_bag[i] == 1 or exhaust_bag[i] == 2) else 0 for i in range(len(p2_data))]
        udds2_w =  [p2_data[i] if (exhaust_bag[i] == 4 or exhaust_bag[i] == 5) else 0 for i in range(len(p2_data))]
        highway_w = [p2_data[i] if exhaust_bag[i] == 3 else 0 for i in range(len(p2_data))]
        us06_w = [p2_data[i] if (exhaust_bag[i] == 6 or exhaust_bag[i] == 7) else 0 for i in range(len(p2_data))]
        
        no_cycle_wh = self.fill_cumulative_list(no_cycle)
        udds1_wh = self.fill_cumulative_list(udds1_w)
        udds2_wh = self.fill_cumulative_list(udds2_w)
        highway_wh = self.fill_cumulative_list(highway_w)
        us06_wh = self.fill_cumulative_list(us06_w)

        u_p_no_cycle = [0 if no_cycle[i] == 0 else ((0.0035 * no_cycle[i]) + 54) for i in range(len(p2_data))]
        u_p_no_cycle_percentage = [0 if no_cycle[i] == 0 else (u_p_no_cycle[i] / no_cycle[i]) for i in range(len(p2_data))]
        u_p_udds1 = [0 if udds1_w[i] == 0 else ((0.0035 * udds1_w[i]) + 54) for i in range(len(p2_data))]
        u_p_udds1_percentage = [0 if udds1_w[i] == 0 else (u_p_udds1[i] / udds1_w[i]) for i in range(len(p2_data))]
        u_p_udds2 = [0 if udds2_w[i] == 0 else ((0.0035 * udds2_w[i]) + 54) for i in range(len(p2_data))]
        u_p_udds2_percentage = [0 if udds2_w[i] == 0 else (u_p_udds2[i] / udds2_w[i]) for i in range(len(p2_data))]
        u_p_highway = [0 if highway_w[i] == 0 else ((0.0035 * highway_w[i]) + 54) for i in range(len(p2_data))]
        u_p_highway_percentage = [0 if highway_w[i] == 0 else (u_p_highway[i] / highway_w[i]) for i in range(len(p2_data))]
        u_p_us06 = [0 if us06_w[i] == 0 else ((0.0035 * us06_w[i]) + 54) for i in range(len(p2_data))]
        u_p_us06_percentage = [0 if us06_w[i] == 0 else (u_p_us06[i] / us06_w[i]) for i in range(len(p2_data))]
        
        # Prepare a DataFrame with the values for easy export to Excel
        group_channel_dataframe = pd.DataFrame({
            "DAQ_Time[s]": daq_time,
            "P2": p2_data,
            "Exhaust_Bag": exhaust_bag,
            "No_cycle": no_cycle,
            "UDDS1_[W]": udds1_w,
            "UDDS2_[W]": udds2_w,
            "Highway_[W]": highway_w,
            "US06_[W]": us06_w,
            "No_cycle_[Wh]": no_cycle_wh,
            "UDDS1_[Wh]": udds1_wh,
            "UDDS2_[Wh]": udds2_wh,
            "Highway_[Wh]": highway_wh,
            "US06_[Wh]": us06_wh,
            "u(P)_no_cycle": u_p_no_cycle,  
            "u(P)_no_cycle_[%]": u_p_no_cycle_percentage,
            "u(P)_UDDS1": u_p_udds1,  
            "u(P)_UDDS1_[%]": u_p_udds1_percentage,
            "u(P)_UDDS2": u_p_udds2,  
            "u(P)_UDDS2_[%]": u_p_udds2_percentage,
            "u(P)_Highway": u_p_highway,  
            "u(P)_Highway_[%]": u_p_highway_percentage,
            "u(P)_US06": u_p_us06,  
            "u(P)_US06_[%]": u_p_us06_percentage   
        })       
            
        return group_channel_dataframe
    
    def get_summary_table(self, group_channel_dataframe):

        energy_channels = ['No_cycle_[Wh]', 'UDDS1_[Wh]', 'UDDS2_[Wh]', 'Highway_[Wh]', 'US06_[Wh]']
        u_energy_channels = ['u(P)_no_cycle', 'u(P)_UDDS1', 'u(P)_UDDS2', 'u(P)_Highway', 'u(P)_US06']
        energy_values = [group_channel_dataframe[channel].dropna().iloc[-1] for channel in energy_channels]
        u_energy_values = [(group_channel_dataframe[channel].dropna().sum())*0.1/3600 for channel in u_energy_channels]
        u_energy_percent = [(u_energy_values[i] / energy_values[i]) * 100 for i in range(len(u_energy_values))]
        # u_energy_sqrt = [(np.sqrt(group_channel_dataframe[channel].dropna()).sum())*0.1/3600 for channel in u_energy_channels]
        u_energy_sqrt = [np.sqrt(np.sum(group_channel_dataframe[channel]**2)) * 0.1 / 3600 for channel in u_energy_channels]
        u_energy_sqrt_percent =[(u_energy_sqrt[i] / energy_values[i]) * 100 for i in range(len(u_energy_values))]
        
        summary_data = [
            ["SUMMARY (cycle totals)", "No-cycle", "UDDS 1", "UDDS 2", "Highway", "US06", "Total"],
            ["Energy [Wh]", energy_values[0], energy_values[1], energy_values[2], energy_values[3], energy_values[4], sum(energy_values)],
            ["u (Energy)", u_energy_values[0], u_energy_values[1], u_energy_values[2], u_energy_values[3], u_energy_values[4], (sum(u_energy_values)-u_energy_values[0])],
            ["u (Energy) [%]", u_energy_percent[0], u_energy_percent[1], u_energy_percent[2], u_energy_percent[3], u_energy_percent[4], ((sum(u_energy_values)-u_energy_values[0]) / sum(energy_values))* 100 ],
            ["u_sqrt (Energy)", u_energy_sqrt[0], u_energy_sqrt[1], u_energy_sqrt[2], u_energy_sqrt[3], u_energy_sqrt[4], (sum(u_energy_sqrt) - u_energy_sqrt[0])],
            ["u_sqrt (Energy) [%]", u_energy_sqrt_percent[0], u_energy_sqrt_percent[1], u_energy_sqrt_percent[2], u_energy_sqrt_percent[3], u_energy_sqrt_percent[4], ((sum(u_energy_sqrt) - u_energy_sqrt[0]) / sum(energy_values))* 100 ]
        ]

        return summary_data

    def add_summary_table(self, sheet_name, summary_data):
        # Load the workbook and create a new sheet if it doesn't exist
        wb = openpyxl.load_workbook(self.output_file_path)
        
        # Check if sheet exists, if not, create it
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(sheet_name)
        else:
            sheet = wb[sheet_name]

        # Write the summary data to the sheet, starting from row 1
        for row_idx, row_data in enumerate(summary_data, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save workbook after writing the summary
        wb.save(self.output_file_path)
        wb.close()

    def manage_depletion_tdms_file(self, test_ID_list):
        test_ID_list = [62005017]
        # Loop through each test ID in the list
        for test_id in test_ID_list:
            
            self.tdms_file_path = self.tdms_data_directory + f"/{test_id} Test Data.tdms"
            
            print(self.tdms_file_path)
            tdms_file = TdmsFile.read(self.tdms_file_path, memmap_dir=None)

            # Get the DataFrame and summary data
            group_channel_dataframe = self.get_depletion_data_group_channel_dataframe(tdms_file)
            summary_data = self.get_depletion_summary_table(group_channel_dataframe)
            
            # First, write the summary table to the sheet
            sheet_name = "wh_cal_" + str(test_id)
            self.add_summary_table(sheet_name, summary_data)
            
            # Write the DataFrame below the summary table using pd.ExcelWriter
            with pd.ExcelWriter(self.output_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                # Write the DataFrame to the specified sheet, below the summary table
                group_channel_dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(summary_data) + 2)

            # Load workbook to apply formatting
            wb = openpyxl.load_workbook(self.output_file_path)
            
            # Apply borders and bold column headers to the summary table
            self.style_dataframe(wb, sheet_name, start_row=1, dataframe=pd.DataFrame(summary_data))

            # Apply borders and bold column headers after writing the DataFrame
            self.style_dataframe(wb, sheet_name, start_row=len(summary_data) + 2, dataframe=group_channel_dataframe)

            # Save and close the workbook
            wb.save(self.output_file_path)
            wb.close()
            
            print(f"Data successfully written to sheet '{sheet_name}' in {self.output_file_path}")


    def get_depletion_data_group_channel_dataframe(self, tdms_file):
            
        group_channel_dataframe = pd.DataFrame()
   
        # Access the 'Data' group
        group_data = tdms_file["Data"]

        # Read the DAQ_Time[s] and P2 channels
        daq_time = group_data["DAQ_Time[s]"].data
        pwr_w = group_data["P2"].data
        eng_wh = self.fill_cumulative_list(pwr_w)
        u_p = [0 if pwr_w[i] == 0 else ((0.35/100 * pwr_w[i]) + (0.09/100*60000)) for i in range(len(pwr_w))]
        u_p_percentage = [0 if pwr_w[i] == 0 else (u_p[i] / pwr_w[i]) for i in range(len(pwr_w))]
        pwr_min = [(pwr_w[i] - u_p[i]) for i in range(len(pwr_w))]
        pwr_max = [(pwr_w[i] + u_p[i]) for i in range(len(pwr_w))]
        eng_min = self.fill_cumulative_list(pwr_min)
        eng_max = self.fill_cumulative_list(pwr_max)
        
        # Prepare a DataFrame with the values for easy export to Excel
        group_channel_dataframe = pd.DataFrame({
            "DAQ_Time[s]": daq_time,
            "PWR[W]": pwr_w,
            "Energy_[Wh]": eng_wh,
            "u(P)": u_p,  
            "u(P)_[%]": u_p_percentage,
            "PWR_min": pwr_min,  
            "PWR_max": pwr_max,
            "ENG_min": eng_min,  
            "ENG_max": eng_max,   
        })

        return group_channel_dataframe
    

    def get_depletion_summary_table(self, group_channel_dataframe):
        
        energy_values = group_channel_dataframe['Energy_[Wh]'].dropna().iloc[-1]
        u_energy_values = (group_channel_dataframe['u(P)'].sum()) * 0.1 / 3600
        u_energy_percent = (u_energy_values / energy_values) * 100 
        u_energy_sqrt = np.sqrt(np.sum(group_channel_dataframe['u(P)']**2)) * 0.1 / 3600
        u_energy_sqrt_percent = (u_energy_sqrt / energy_values) * 100 

        summary_data = [
            ["SUMMARY ", "No-cycle"],
            ["Energy [Wh]", energy_values],
            ["u (Energy)", u_energy_values],
            ["u (Energy) [%]", u_energy_percent],
            ["u_sqrt (Energy)", u_energy_sqrt],
            ["u_sqrt (Energy) [%]", u_energy_sqrt_percent]
        ]

        return summary_data
    
    def manage_categorial_summary(self):
        """
        create a summary sheet based on each category
        summary sheet will contain summary table and different plots
        """
        
    def style_dataframe(self, wb, sheet_name, start_row, dataframe):
        # Load the specified sheet
        sheet = wb[sheet_name]

        # Define border style
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Bold font for column headers
        bold_font = Font(bold=True)

        # Get the number of rows and columns in the DataFrame
        n_rows, n_cols = dataframe.shape
        
        # Apply styles to the column headers (first row of the DataFrame or summary)
        for col_idx in range(1, n_cols + 1):
            cell = sheet.cell(row=start_row, column=col_idx)
            cell.font = bold_font  # Bold the column header
            cell.border = thin_border  # Apply border to the column header
        
        # Apply borders only to the range where the DataFrame has actual data
        for row_idx in range(start_row + 1, start_row + n_rows + 1):
            for col_idx in range(1, n_cols + 1):
                # Get the actual value from the DataFrame for this cell
                cell_value = dataframe.iloc[row_idx - start_row - 1, col_idx - 1]
                
                # Only apply border if the cell has non-NaN, non-empty data
                if pd.notna(cell_value) and cell_value != "":
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border  # Apply border to each non-empty cell

    def __del__(self):
        print(f"{self.object_name} is destroyed.")