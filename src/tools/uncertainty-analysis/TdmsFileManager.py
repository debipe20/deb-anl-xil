"""
**********************************************************************************

TdmsFileManager.py
Created by: Debashis Das
Argonne National Laboratory
Transportation and Power Systems Division

**********************************************************************************
  
Description:
------------
This script defines the TdmsFileManager class, which handles the processing, 
analysis, and visualization of TDMS test data. Key features include reading 
TDMS files, summarizing energy data, calculating uncertainty percentages, 
and writing formatted results to Excel workbooks.

Methods available in the class:
- initialize_ube_fre_variables(): Initialize variables related to UBE and FRE metrics.
- set_test_id_list(): Assigns test ID lists to categories and triggers analysis.
- conduct_mct_test_analysis(): Conducts analysis for non-empty test ID categories.
- manage_mct_test_analysis(): Manages the processing and summarization of TDMS files.
- get_sequence_group_channel_dataframe(): Creates a DataFrame for energy data sequences.
- get_depletion_and_charge_data_group_channel_dataframe(): Creates DataFrame for depletion/charge cycles.
- get_sequence_summary_table(): Summarizes sequence-based energy data into a structured format.
- get_depletion_and_charge_summary_table(): Summarizes depletion/charge cycle data.
- write_test_data(): Writes processed data and summaries to an Excel file.
- fill_cumulative_list(): Computes cumulative energy values for a list.
- add_summary_table(): Adds a summary table to an Excel sheet.
- manage_categorial_summary(): Handles the creation of summary sheets by category.
- style_dataframe(): Applies formatting styles to DataFrame content in an Excel sheet.

"""

import numpy as np
import openpyxl
import pandas as pd
from nptdms import TdmsFile
from openpyxl.styles import Border, Side, Font, Alignment
from PlotManager import PlotManager

class TdmsFileManager:
    def __init__(self, output_file_path, plot_directory, tdms_directory, accuracy_parameter1, accuracy_parameter2, accuracy_parameter3):
        """
        Initializes the TdmsFileManager with file paths and parameters.

        Args:
            output_file_path (str): Path to the Excel file for saving results.
            plot_directory (str): Directory path for saving plots.
            tdms_directory (str): Directory path where TDMS files are stored.
            accuracy_parameter1 (float): Accuracy parameter 1 for uncertainty calculations.
            accuracy_parameter2 (float): Accuracy parameter 2 for uncertainty calculations.
            accuracy_parameter3 (float): Accuracy parameter 3 for uncertainty calculations.
        """
        self.output_file_path = output_file_path
        self.plot_directory = plot_directory
        self.tdms_data_directory = tdms_directory
        self.accuracy_parameter1 = accuracy_parameter1
        self.accuracy_parameter2 = accuracy_parameter2
        self.accuracy_parameter3 = accuracy_parameter3

        self.initialize_ube_fre_variables()

    def initialize_ube_fre_variables(self):
        """
        Initializes UBE and FRE variables for energy metrics.
        """
        self.ube_wh, self.u_ube_wh, self.u_ube_rms, self.u_ube_sq = 0, 0, 0, 0
        self.fre_wh, self.u_fre_wh, self.u_fre_percent, self.u_fre_rms, self.u_fre_rms_percent = 0, 0, 0, 0, 0
        self.bar_chart_test_id_list = []

    def set_test_id_list(self, test_id_list1, test_id_list2, test_id_list3, test_id_list4, test_id_list5):
        """
        Sets test ID lists for various categories and initiates MCT test analysis.

        Args:
            test_id_list1, test_id_list2, test_id_list3, test_id_list4, test_id_list5 (list): 
                Lists of test IDs for corresponding categories.
        """
        self.test_id_list_category_1st, self.test_id_list_category_2nd, self.test_id_list_category_3rd, self.test_id_list_category_4th, self.test_id_list_category_5th = test_id_list1, test_id_list2, test_id_list3, test_id_list4, test_id_list5

        if self.test_id_list_category_1st:
            self.manage_mct_test(self.test_id_list_category_1st)

    def conduct_mct_test_analysis(self, test_id_list1, test_id_list2, test_id_list3, test_id_list4, test_id_list5):
        """
        Conducts MCT test analysis for non-empty test ID lists.

        Args:
            test_id_list1, test_id_list2, test_id_list3, test_id_list4, test_id_list5 (list): 
                Lists of test IDs for corresponding categories.
        """
        # Assign the lists to instance variables
        self.test_id_list_category_1st = test_id_list1
        self.test_id_list_category_2nd = test_id_list2
        self.test_id_list_category_3rd = test_id_list3
        self.test_id_list_category_4th = test_id_list4
        self.test_id_list_category_5th = test_id_list5

        # Create a dictionary to associate category lists with names (for better debugging)
        test_id_lists = {
            "Category 1": self.test_id_list_category_1st,
            "Category 2": self.test_id_list_category_2nd,
            "Category 3": self.test_id_list_category_3rd,
            "Category 4": self.test_id_list_category_4th,
            "Category 5": self.test_id_list_category_5th
        }

        # Iterate through the lists and call manage_mct_test() for non-empty ones
        for category_name, test_id_list in test_id_lists.items():
            if test_id_list:  # Check if the list is not empty
                print(f"Processing {category_name} with Test IDs: {test_id_list}")
                self.manage_mct_test_analysis(test_id_list)

        
    def manage_mct_test_analysis(self, test_id_list):
        """
        Manages the processing, analysis, and summarization of TDMS files.

        Args:
            test_id_list (list): List of test IDs to process.
        """
        group_channel_dataframe_sequence1, group_channel_dataframe_depletion1, group_channel_dataframe_sequence2, group_channel_dataframe_depletion2, group_channel_dataframe_charge = (pd.DataFrame() for i in range(5))
        summary_data_sequence1, summary_data_depletion1, summary_data_sequence2, summary_data_depletion2, summary_data_charge = ([] for i in range(5))
        summary_title_list = []
        
        for test_id in test_id_list:
            
            self.tdms_file_path = self.tdms_data_directory + f"/{test_id} Test Data.tdms"
            print(f"Processing '{self.tdms_file_path}' TDMS file")
            tdms_file = TdmsFile.read(self.tdms_file_path, memmap_dir=None)
             
            test_id_index = test_id_list.index(test_id)
            summary_title_list.append("Summary Data of Test ID " + str(test_id))
            
            if test_id_index == 0:
                group_channel_dataframe_sequence1 = self.get_sequence_group_channel_dataframe(tdms_file, test_id_index)
                summary_data_sequence1 = self.get_sequence_summary_table(group_channel_dataframe_sequence1, test_id_index)
                self.write_test_data(test_id, group_channel_dataframe_sequence1, summary_data_sequence1)
                self.bar_chart_test_id_list.append(str(test_id))
                
            elif test_id_index == 1:
                group_channel_dataframe_depletion1 = self.get_depletion_and_charge_data_group_channel_dataframe(tdms_file, test_id_index)
                summary_data_depletion1 = self.get_depletion_and_charge_summary_table(group_channel_dataframe_depletion1, test_id_index)
                self.write_test_data(test_id, group_channel_dataframe_depletion1, summary_data_depletion1)
                
            elif test_id_index == 2:
                group_channel_dataframe_sequence2 = self.get_sequence_group_channel_dataframe(tdms_file, test_id_index)
                summary_data_sequence2 = self.get_sequence_summary_table(group_channel_dataframe_sequence2, test_id_index)
                self.write_test_data(test_id, group_channel_dataframe_sequence2, summary_data_sequence2)
                self.bar_chart_test_id_list.append(str(test_id))
                
            elif test_id_index == 3:
                group_channel_dataframe_depletion2 = self.get_depletion_and_charge_data_group_channel_dataframe(tdms_file, test_id_index)
                summary_data_depletion2 = self.get_depletion_and_charge_summary_table(group_channel_dataframe_depletion2, test_id_index)
                self.write_test_data(test_id, group_channel_dataframe_depletion2, summary_data_depletion2)

            elif test_id_index == 4:
                group_channel_dataframe_charge = self.get_depletion_and_charge_data_group_channel_dataframe(tdms_file, test_id_index)
                summary_data_charge = self.get_depletion_and_charge_summary_table(group_channel_dataframe_charge, test_id_index)
                self.write_test_data(test_id, group_channel_dataframe_charge, summary_data_charge)
                self.manage_categorial_summary(test_id_list[0], 0, summary_data_sequence1, summary_title_list[0])
                self.manage_categorial_summary(test_id_list[1], 1, summary_data_depletion1, summary_title_list[1])
                self.manage_categorial_summary(test_id_list[2], 2, summary_data_sequence2, summary_title_list[2])
                self.manage_categorial_summary(test_id_list[3], 3, summary_data_depletion2, summary_title_list[3])
                self.manage_categorial_summary(test_id_list[4], test_id_index, summary_data_charge, summary_title_list[4])

                plot_manager = PlotManager(summary_data_sequence1, summary_data_sequence2, self.bar_chart_test_id_list, self.output_file_path, self.plot_directory, self.categorial_summary_sheet_name)
                plot_manager.plot_uncertainty_percentage()
                plot_manager.plot_energy_analysis()
                del plot_manager

    def get_sequence_group_channel_dataframe(self, tdms_file, test_id_index):
        """
        Creates a DataFrame containing energy data for different drive cycles.

        Args:
            tdms_file (TdmsFile): TDMS file object containing test data.
            test_id_index (int): Index of the test ID in the category list.

        Returns:
            pd.DataFrame: DataFrame containing energy and uncertainty data for the specified test ID.
        """    
        group_channel_dataframe = pd.DataFrame()
        no_cycle_wh, udds1_wh, udds2_wh, highway_wh, us06_wh = ([] for i in range(5))
        # Access the 'Data' group
        group_data = tdms_file["Data"]

        # Read the DAQ_Time[s] and P2 channels
        daq_time = group_data["DAQ_Time[s]"].data
        p2_data = group_data["P2"].data
        exhaust_bag = group_data["Exhaust_Bag"].data
        time_values = [i * 0.1 for i in range(len(daq_time))]

        if test_id_index == 0:
            no_cycle = [p2_data[i] if exhaust_bag[i] == 0 else 0 for i in range(len(p2_data))]
            udds1_w =  [p2_data[i] if (exhaust_bag[i] == 1 or exhaust_bag[i] == 2) else 0 for i in range(len(p2_data))]
            highway_w = [p2_data[i] if exhaust_bag[i] == 3 else 0 for i in range(len(p2_data))]
            udds2_w =  [p2_data[i] if (exhaust_bag[i] == 4 or exhaust_bag[i] == 5) else 0 for i in range(len(p2_data))]            
            us06_w = [p2_data[i] if (exhaust_bag[i] == 6 or exhaust_bag[i] == 7) else 0 for i in range(len(p2_data))]
            
        elif test_id_index == 2:
            no_cycle = [p2_data[i] if exhaust_bag[i] == 0 else 0 for i in range(len(p2_data))]
            us06_w = [p2_data[i] if (exhaust_bag[i] == 1 or exhaust_bag[i] == 2) else 0 for i in range(len(p2_data))]
            udds1_w =  [p2_data[i] if (exhaust_bag[i] == 3 or exhaust_bag[i] == 4) else 0 for i in range(len(p2_data))]
            highway_w = [p2_data[i] if exhaust_bag[i] == 5 else 0 for i in range(len(p2_data))]
            udds2_w =  [p2_data[i] if (exhaust_bag[i] == 6 or exhaust_bag[i] == 7) else 0 for i in range(len(p2_data))]            
        
        no_cycle_wh = self.fill_cumulative_list(no_cycle)
        udds1_wh = self.fill_cumulative_list(udds1_w)
        udds2_wh = self.fill_cumulative_list(udds2_w)
        highway_wh = self.fill_cumulative_list(highway_w)
        us06_wh = self.fill_cumulative_list(us06_w)

        u_p_no_cycle = [0 if no_cycle[i] == 0 else abs((((self.accuracy_parameter1 / 100) * no_cycle[i]) + ((self.accuracy_parameter2 / 100) * 60000))) for i in range(len(p2_data))]
        u_p_no_cycle_percentage = [0 if no_cycle[i] == 0 else (u_p_no_cycle[i] / no_cycle[i]) for i in range(len(p2_data))]
        u_p_udds1 = [0 if udds1_w[i] == 0 else abs((((self.accuracy_parameter1 / 100) * udds1_w[i]) + ((self.accuracy_parameter2 / 100) * 60000))) for i in range(len(p2_data))]
        u_p_udds1_percentage = [0 if udds1_w[i] == 0 else (u_p_udds1[i] / udds1_w[i]) for i in range(len(p2_data))]
        u_p_udds2 = [0 if udds2_w[i] == 0 else abs((((self.accuracy_parameter1 / 100) * udds2_w[i]) + ((self.accuracy_parameter2 / 100) * 60000))) for i in range(len(p2_data))]
        u_p_udds2_percentage = [0 if udds2_w[i] == 0 else (u_p_udds2[i] / udds2_w[i]) for i in range(len(p2_data))]
        u_p_highway = [0 if highway_w[i] == 0 else abs((((self.accuracy_parameter1 / 100) * highway_w[i]) + ((self.accuracy_parameter2 / 100) * 60000))) for i in range(len(p2_data))]
        u_p_highway_percentage = [0 if highway_w[i] == 0 else (u_p_highway[i] / highway_w[i]) for i in range(len(p2_data))]
        u_p_us06 = [0 if us06_w[i] == 0 else abs((((self.accuracy_parameter1 / 100) * us06_w[i]) + ((self.accuracy_parameter2 / 100) * 60000))) for i in range(len(p2_data))]
        u_p_us06_percentage = [0 if us06_w[i] == 0 else (u_p_us06[i] / us06_w[i]) for i in range(len(p2_data))]
    
        # Prepare a DataFrame with the values for easy export to Excel
        if test_id_index == 0:
            group_channel_dataframe = pd.DataFrame({
                "DAQ_Time[s]": daq_time,
                "Time[s]": time_values,
                "P2": p2_data,
                "Exhaust_Bag": exhaust_bag,
                "No_cycle": no_cycle,
                "UDDS1_[W]": udds1_w,
                "Highway_[W]": highway_w,
                "UDDS2_[W]": udds2_w,            
                "US06_[W]": us06_w,
                "No_cycle_[Wh]": no_cycle_wh,
                "UDDS1_[Wh]": udds1_wh,
                "UDDS2_[Wh]": udds2_wh,
                "Highway_[Wh]": highway_wh,
                "US06_[Wh]": us06_wh,
                "u(P)_no_cycle": u_p_no_cycle,  
                "u(P)_no_cycle_[%]": u_p_no_cycle_percentage,
                "u(P)_UDDS1": u_p_udds1,  
                "u(P)_UDDS1_[%]": u_p_udds1_percentage,
                "u(P)_UDDS2": u_p_udds2,  
                "u(P)_UDDS2_[%]": u_p_udds2_percentage,
                "u(P)_Highway": u_p_highway,  
                "u(P)_Highway_[%]": u_p_highway_percentage,
                "u(P)_US06": u_p_us06,  
                "u(P)_US06_[%]": u_p_us06_percentage   
            }) 
            
        elif test_id_index == 2:
            group_channel_dataframe = pd.DataFrame({
            "DAQ_Time[s]": daq_time,
            "Time[s]": time_values,
            "P2": p2_data,
            "Exhaust_Bag": exhaust_bag,
            "No_cycle": no_cycle,
            "US06_[W]": us06_w,
            "UDDS1_[W]": udds1_w,
            "Highway_[W]": highway_w,
            "UDDS2_[W]": udds2_w,                    
            "No_cycle_[Wh]": no_cycle_wh,
            "US06_[Wh]": us06_wh,
            "UDDS1_[Wh]": udds1_wh,
            "Highway_[Wh]": highway_wh,
            "UDDS2_[Wh]": udds2_wh,        
            "u(P)_no_cycle": u_p_no_cycle,  
            "u(P)_no_cycle_[%]": u_p_no_cycle_percentage,
            "u(P)_US06": u_p_us06,  
            "u(P)_US06_[%]": u_p_us06_percentage,
            "u(P)_UDDS1": u_p_udds1,  
            "u(P)_UDDS1_[%]": u_p_udds1_percentage,
            "u(P)_Highway": u_p_highway,  
            "u(P)_Highway_[%]": u_p_highway_percentage,
            "u(P)_UDDS2": u_p_udds2,  
            "u(P)_UDDS2_[%]": u_p_udds2_percentage           
        })
              
        return group_channel_dataframe
    
    def get_depletion_and_charge_data_group_channel_dataframe(self, tdms_file, test_id_index):
        """
        Creates a DataFrame containing energy depletion and charge cycle data.

        Args:
            tdms_file (TdmsFile): TDMS file object containing test data.
            test_id_index (int): Index of the test ID in the category list.

        Returns:
            pd.DataFrame: DataFrame containing energy and uncertainty data for depletion and charge cycles.
        """            
        group_channel_dataframe = pd.DataFrame()
   
        # Access the 'Data' group
        group_data = tdms_file["Data"]

        # Read the DAQ_Time[s] and P2 channels
        daq_time = group_data["DAQ_Time[s]"].data
        time_values = [i * 0.1 for i in range(len(daq_time))]
        
        if test_id_index == 1:
            pwr_w = group_data["P2"].data
            eng_wh = self.fill_cumulative_list(pwr_w)
            u_p = [((self.accuracy_parameter1/100 * pwr_w[i]) + (self.accuracy_parameter2 / 100*60000)) for i in range(len(pwr_w))]

        elif test_id_index == 3:
            pwr_w = group_data["P1"].data
            eng_wh = self.fill_cumulative_list(pwr_w)
            u_p = [((self.accuracy_parameter1/100 * pwr_w[i]) + (self.accuracy_parameter2 / 100*60000)) for i in range(len(pwr_w))]
            
        elif test_id_index == 4: 
            pwr_w = group_data["P9"].data
            eng_wh = self.fill_cumulative_list(pwr_w)
            u_p = [((self.accuracy_parameter1/100 * pwr_w[i]) + (self.accuracy_parameter3 / 100*60000)) for i in range(len(pwr_w))]

        u_p_percentage = [0 if pwr_w[i] == 0 else (u_p[i] / pwr_w[i]) for i in range(len(pwr_w))]
        # pwr_min = [(pwr_w[i] - u_p[i]) for i in range(len(pwr_w))]
        # pwr_max = [(pwr_w[i] + u_p[i]) for i in range(len(pwr_w))]
        # eng_min = self.fill_cumulative_list(pwr_min)
        # eng_max = self.fill_cumulative_list(pwr_max)
        
        # Prepare a DataFrame with the values for easy export to Excel
        group_channel_dataframe = pd.DataFrame({
            "DAQ_Time[s]": daq_time,
            "Time[s]": time_values,
            "PWR[W]": pwr_w,
            "Energy_[Wh]": eng_wh,
            "u(P)": u_p,  
            "u(P)_[%]": u_p_percentage,
            # "PWR_min": pwr_min,  
            # "PWR_max": pwr_max,
            # "ENG_min": eng_min,  
            # "ENG_max": eng_max,   
        })

        return group_channel_dataframe
    
    # def get_sequence_summary_table(self, group_channel_dataframe, test_id_index):

    #     if test_id_index == 0:
    #         energy_channels = ['No_cycle_[Wh]', 'UDDS1_[Wh]', 'UDDS2_[Wh]', 'Highway_[Wh]', 'US06_[Wh]']
    #         u_energy_channels = ['u(P)_no_cycle', 'u(P)_UDDS1', 'u(P)_UDDS2', 'u(P)_Highway', 'u(P)_US06']

    #     elif test_id_index == 2:
    #         energy_channels = ['No_cycle_[Wh]', 'US06_[Wh]', 'UDDS1_[Wh]', 'Highway_[Wh]', 'UDDS2_[Wh]']
    #         u_energy_channels = ['u(P)_no_cycle', 'u(P)_US06', 'u(P)_UDDS1', 'u(P)_Highway', 'u(P)_UDDS2']
        
    #     else:
    #         raise ValueError(f"Invalid test_id_index: {test_id_index}")

    #     energy_values = [group_channel_dataframe[channel].dropna().iloc[-1] for channel in energy_channels]
    #     # u_energy_values = [(group_channel_dataframe[channel].dropna().sum()) * 0.1 / 3600 for channel in u_energy_channels]
    #     u_energy_values = [np.sum(group_channel_dataframe[channel]) * 0.1 / 3600 for channel in u_energy_channels]
    #     u_energy_percent = [(u_energy_values[i] / energy_values[i]) * 100 for i in range(len(u_energy_values))]
    #     u_energy_sqrt = [np.sqrt(np.sum(group_channel_dataframe[channel]**2)) * 0.1 / 3600 for channel in u_energy_channels]
    #     u_energy_sqrt_percent =[(u_energy_sqrt[i] / energy_values[i]) * 100 for i in range(len(u_energy_values))]
        
    #     if test_id_index == 0:
    #         summary_data = [
    #             ["SUMMARY (cycle totals)", "No-cycle", "UDDS 1", "UDDS 2", "Highway", "US06", "Total"],
    #             ["Energy [Wh]", energy_values[0], energy_values[1], energy_values[2], energy_values[3], energy_values[4], sum(energy_values)],
    #             ["u (Energy)", u_energy_values[0], u_energy_values[1], u_energy_values[2], u_energy_values[3], u_energy_values[4], (sum(u_energy_values)-u_energy_values[0])],
    #             ["u (Energy) [%]", u_energy_percent[0], u_energy_percent[1], u_energy_percent[2], u_energy_percent[3], u_energy_percent[4], ((sum(u_energy_values)-u_energy_values[0]) / sum(energy_values))* 100],
    #             ["u_sqrt (Energy)", u_energy_sqrt[0], u_energy_sqrt[1], u_energy_sqrt[2], u_energy_sqrt[3], u_energy_sqrt[4], (sum(u_energy_sqrt) - u_energy_sqrt[0])],
    #             ["u_sqrt (Energy) [%]", u_energy_sqrt_percent[0], u_energy_sqrt_percent[1], u_energy_sqrt_percent[2], u_energy_sqrt_percent[3], u_energy_sqrt_percent[4], ((sum(u_energy_sqrt) - u_energy_sqrt[0]) / sum(energy_values))* 100 ]
    #         ]

    #     elif test_id_index == 2:
    #         summary_data = [
    #             ["SUMMARY (cycle totals)", "No-cycle", "US06", "UDDS 1", "Highway", "UDDS 2", "Total"],
    #             ["Energy [Wh]", energy_values[0], energy_values[1], energy_values[2], energy_values[3], energy_values[4], sum(energy_values)],
    #             ["u (Energy)", u_energy_values[0], u_energy_values[1], u_energy_values[2], u_energy_values[3], u_energy_values[4], (sum(u_energy_values) - u_energy_values[0])],
    #             ["u (Energy) [%]", u_energy_percent[0], u_energy_percent[1], u_energy_percent[2], u_energy_percent[3], u_energy_percent[4], ((sum(u_energy_values) - u_energy_values[0]) / sum(energy_values))* 100],
    #             ["u_sqrt (Energy)", u_energy_sqrt[0], u_energy_sqrt[1], u_energy_sqrt[2], u_energy_sqrt[3], u_energy_sqrt[4], (sum(u_energy_sqrt) - u_energy_sqrt[0])],
    #             ["u_sqrt (Energy) [%]", u_energy_sqrt_percent[0], u_energy_sqrt_percent[1], u_energy_sqrt_percent[2], u_energy_sqrt_percent[3], u_energy_sqrt_percent[4], ((sum(u_energy_sqrt) - u_energy_sqrt[0]) / sum(energy_values))* 100 ]
    #         ]           

    #     self.ube_wh = self.ube_wh + sum(energy_values)
    #     self.u_ube_wh = self.u_ube_wh + (sum(u_energy_values) - u_energy_values[0])
    #     self.u_ube_sq = self.u_ube_sq + (sum(u_energy_sqrt) - u_energy_sqrt[0]) * (sum(u_energy_sqrt) - u_energy_sqrt[0])

    #     return summary_data

    def get_sequence_summary_table(self, group_channel_dataframe, test_id_index):
        """
        Summarizes sequence-based energy data into a structured format.

        Args:
            group_channel_dataframe (pd.DataFrame): DataFrame containing sequence energy data.
            test_id_index (int): Index of the test ID in the category list.

        Returns:
            list: Summary data for the sequence-based energy metrics.
        """
        energy_channels = ['No_cycle_[Wh]', 'UDDS1_[Wh]', 'UDDS2_[Wh]', 'Highway_[Wh]', 'US06_[Wh]']
        u_energy_channels = ['u(P)_no_cycle', 'u(P)_UDDS1', 'u(P)_UDDS2', 'u(P)_Highway', 'u(P)_US06']

        energy_values = [group_channel_dataframe[channel].dropna().iloc[-1] for channel in energy_channels]
        # u_energy_values = [(group_channel_dataframe[channel].dropna().sum()) * 0.1 / 3600 for channel in u_energy_channels]
        u_energy_values = [np.sum(group_channel_dataframe[channel]) * 0.1 / 3600 for channel in u_energy_channels]
        u_energy_percent = [(u_energy_values[i] / energy_values[i]) * 100 for i in range(len(u_energy_values))]
        u_energy_sqrt = [np.sqrt(np.sum(group_channel_dataframe[channel]**2)) * 0.1 / 3600 for channel in u_energy_channels]
        u_energy_sqrt_percent =[(u_energy_sqrt[i] / energy_values[i]) * 100 for i in range(len(u_energy_values))]
        
        summary_data = [
            ["SUMMARY (cycle totals)", "No-cycle", "UDDS 1", "UDDS 2", "Highway", "US06", "Total"],
            ["Energy [Wh]", energy_values[0], energy_values[1], energy_values[2], energy_values[3], energy_values[4], sum(energy_values)],
            ["u (Energy)", u_energy_values[0], u_energy_values[1], u_energy_values[2], u_energy_values[3], u_energy_values[4], (sum(u_energy_values)-u_energy_values[0])],
            ["u (Energy) [%]", u_energy_percent[0], u_energy_percent[1], u_energy_percent[2], u_energy_percent[3], u_energy_percent[4], ((sum(u_energy_values)-u_energy_values[0]) / sum(energy_values))* 100],
            ["u_sqrt (Energy)", u_energy_sqrt[0], u_energy_sqrt[1], u_energy_sqrt[2], u_energy_sqrt[3], u_energy_sqrt[4], (sum(u_energy_sqrt) - u_energy_sqrt[0])],
            ["u_sqrt (Energy) [%]", u_energy_sqrt_percent[0], u_energy_sqrt_percent[1], u_energy_sqrt_percent[2], u_energy_sqrt_percent[3], u_energy_sqrt_percent[4], ((sum(u_energy_sqrt) - u_energy_sqrt[0]) / sum(energy_values))* 100 ]
        ]

        self.ube_wh = self.ube_wh + sum(energy_values)
        self.u_ube_wh = self.u_ube_wh + (sum(u_energy_values) - u_energy_values[0])
        self.u_ube_sq = self.u_ube_sq + (sum(u_energy_sqrt) - u_energy_sqrt[0]) * (sum(u_energy_sqrt) - u_energy_sqrt[0])

        return summary_data

    def get_depletion_and_charge_summary_table(self, group_channel_dataframe, test_id_index):
        """
        Summarizes depletion and charge cycle data into a structured format.

        Args:
            group_channel_dataframe (pd.DataFrame): DataFrame containing depletion/charge energy data.
            test_id_index (int): Index of the test ID in the category list.

        Returns:
            list: Summary data for depletion and charge cycle metrics.
        """
        energy_values = group_channel_dataframe['Energy_[Wh]'].dropna().iloc[-1]
        u_energy_values = (group_channel_dataframe['u(P)'].sum()) * 0.1 / 3600
        u_energy_percent = (u_energy_values / energy_values) * 100 
        u_energy_sqrt = np.sqrt(np.sum(group_channel_dataframe['u(P)']**2)) * 0.1 / 3600
        u_energy_sqrt_percent = (u_energy_sqrt / energy_values) * 100 

        summary_data = [
            ["SUMMARY ", "No-cycle"],
            ["Energy [Wh]", energy_values],
            ["u (Energy)", u_energy_values],
            ["u (Energy) [%]", u_energy_percent],
            ["u_sqrt (Energy)", u_energy_sqrt],
            ["u_sqrt (Energy) [%]", u_energy_sqrt_percent]
        ]

        if test_id_index < 4:
            self.ube_wh = self.ube_wh + energy_values
            self.u_ube_wh = self.u_ube_wh + u_energy_values
            self.u_ube_sq = self.u_ube_sq + (u_energy_sqrt * u_energy_sqrt)
        
        elif test_id_index == 4:
            self.fre_wh = energy_values
            self.u_fre_wh = u_energy_values
            self.u_fre_percent = u_energy_percent
            self.u_fre_rms = u_energy_sqrt
            self.u_fre_rms_percent = u_energy_sqrt_percent

        return summary_data
        
    def write_test_data(self,test_id, group_channel_dataframe, summary_data):
        """
        Writes test data and summary information to an Excel sheet.

        Args:
            test_id (int): ID of the test.
            group_channel_dataframe (pd.DataFrame): DataFrame containing test data.
            summary_data (list): Summary data for the test.

        Returns:
            None
        """        
        # First, write the summary table to the sheet
        sheet_name = "Wh_Cal_" + str(test_id)
        self.add_summary_table(sheet_name, summary_data)
        
        # Write the DataFrame below the summary table using pd.ExcelWriter
        with pd.ExcelWriter(self.output_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            # Write the DataFrame to the specified sheet, below the summary table
            group_channel_dataframe.to_excel(writer, sheet_name=sheet_name, index=False, startrow=len(summary_data) + 2)

        # Load workbook to apply formatting
        wb = openpyxl.load_workbook(self.output_file_path)
        
        # Apply borders and bold column headers to the summary table
        self.style_dataframe(wb, sheet_name, start_row=1, dataframe=pd.DataFrame(summary_data))

        # Apply borders and bold column headers after writing the DataFrame
        self.style_dataframe(wb, sheet_name, start_row=len(summary_data) + 3, dataframe=group_channel_dataframe)

        # Save and close the workbook
        wb.save(self.output_file_path)
        wb.close()
        
        print(f"Data successfully written to sheet '{sheet_name}' in {self.output_file_path}")

    def fill_cumulative_list(self, source_list):
        """
        Computes cumulative energy values from a source list.

        Args:
            source_list (list): List of energy data points.

        Returns:
            list: Cumulative energy values.
        """
        cumulative_list = []
        
        for i in range(len(source_list)):
            if i == 0:
                cumulative_list.append((source_list[i] * 0.1) / 3600)
            else:
                cumulative_list.append(cumulative_list[i-1] + (source_list[i] * 0.1) / 3600)
        return cumulative_list

    def add_summary_table(self, sheet_name, summary_data , title=None):
        """
        Adds a summary table to a specified Excel sheet.

        Args:
            sheet_name (str): Name of the Excel sheet.
            summary_data (list): Summary data to write.
            title (str, optional): Title of the summary table.

        Returns:
            None
        """
        # Load the workbook and create a new sheet if it doesn't exist
        wb = openpyxl.load_workbook(self.output_file_path)
        
        # Check if sheet exists, if not, create it
        if sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(sheet_name)
        else:
            sheet = wb[sheet_name]

        start_row = 1  # Define starting row
    
        # Add a title or description if provided
        if title:
            sheet.cell(row=start_row, column=1, value=title)
            sheet.cell(row=start_row, column=1).font = Font(bold=True)
            start_row += 1  

        # Write the summary data to the sheet, starting from the next row
        for row_idx, row_data in enumerate(summary_data, start=start_row):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)
        # # Write the summary data to the sheet, starting from row 1
        # for row_idx, row_data in enumerate(summary_data, start=1):
        #     for col_idx, cell_value in enumerate(row_data, start=1):
        #         sheet.cell(row=row_idx, column=col_idx, value=cell_value)

        # Save workbook after writing the summary
        wb.save(self.output_file_path)
        wb.close()
    
    def manage_categorial_summary(self, test_id, test_id_index, summary_data, title=None):
        """
        Creates summary sheets for each category, including plots and metrics.

        Args:
            test_id (int): ID of the test.
            test_id_index (int): Index of the test ID in the category list.
            summary_data (list): Summary data for the test.
            title (str, optional): Title for the summary section.

        Returns:
            None
        """
        ube_summary_data = []
        fre_summary_data = []
        
        if test_id in self.test_id_list_category_1st:
            self.categorial_summary_sheet_name = "1_Summary"

        elif test_id in self.test_id_list_category_2nd:
            self.categorial_summary_sheet_name = "2_Summary"

        elif test_id in self.test_id_list_category_3rd:
            self.categorial_summary_sheet_name = "3_Summary"

        elif test_id in self.test_id_list_category_4th:
            self.categorial_summary_sheet_name = "4_Summary"

        elif test_id in self.test_id_list_category_5th:
            self.categorial_summary_sheet_name = "5_Summary" 

        if test_id_index == 4:
            ube_title =  "UBE Summary of Category1"
            fre_title = "FRE Summary of Category1"
            self.u_ube_rms = np.sqrt(self.u_ube_sq)
            ube_summary_data = [
            ["UBE SUMMARY", "Value"],
            ["UBE [Wh]", self.ube_wh],
            ["u(UBE)", self.u_ube_wh],
            ["u(UBE) [%]", (self.u_ube_wh /self.ube_wh) if self.ube_wh != 0 else None],
            ["u(UBE)_sqrt", self.u_ube_rms],
            ["u(UBE)_sqrt [%]", (self.u_ube_rms / self.ube_wh) if self.ube_wh != 0 else None]
        ]

            fre_summary_data = [
            ["FRE SUMMARY", "Value"],
            ["FRE [Wh]", self.fre_wh],
            ["u(FRE)", self.u_fre_wh],
            ["u(FRE) [%]", self.u_fre_percent],
            ["u(FRE)_sqrt", self.u_fre_rms],
            ["u(FRE)_sqrt [%]", self.u_fre_rms_percent]
        ] 

        wb = openpyxl.load_workbook(self.output_file_path)
        
        # Check if sheet exists, if not, create it
        if self.categorial_summary_sheet_name not in wb.sheetnames:
            sheet = wb.create_sheet(self.categorial_summary_sheet_name)
        else:
            sheet = wb[self.categorial_summary_sheet_name]

        # Find the next available row to append data
        start_row = sheet.max_row + 2 if sheet.max_row > 1 else 1  # Start at row 1 for the first table
        
        if title:
            sheet.cell(row=start_row, column=1, value=title)
            sheet.cell(row=start_row, column=1).font = Font(bold=True)
            start_row += 1

        # Write the summary data to the sheet, starting from the next available row
        for row_idx, row_data in enumerate(summary_data, start=start_row):
            for col_idx, cell_value in enumerate(row_data, start=1):
                sheet.cell(row=row_idx, column=col_idx, value=cell_value)
    
        self.style_dataframe(wb, self.categorial_summary_sheet_name, start_row, dataframe=pd.DataFrame(summary_data))
        
        if test_id_index == 4: 
            # Find the next available row after summary data
            start_row += len(summary_data) + 3  # Leave gap of two rows
            
            sheet.cell(row=start_row, column=1, value=ube_title)
            sheet.cell(row=start_row, column=1).font = Font(bold=True)
            start_row += 1
            
            # Write the ube_summary data to the sheet, starting from the next available row
            for row_idx, row_data in enumerate(ube_summary_data, start=start_row):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            self.style_dataframe(wb, self.categorial_summary_sheet_name, start_row, dataframe=pd.DataFrame(ube_summary_data))
            
            
            start_row += len(ube_summary_data) + 3  # Leave gap of two rows
            
            sheet.cell(row=start_row, column=1, value=fre_title)
            sheet.cell(row=start_row, column=1).font = Font(bold=True)
            start_row += 1
            
            # Write the fre_summary data to the sheet, starting from the next available row
            for row_idx, row_data in enumerate(fre_summary_data, start=start_row):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            
            self.style_dataframe(wb, self.categorial_summary_sheet_name, start_row, dataframe=pd.DataFrame(fre_summary_data))
            self.initialize_ube_fre_variables()

        # Save workbook after writing the summary
        wb.save(self.output_file_path)
        wb.close()
        
    def style_dataframe(self, wb, sheet_name, start_row, dataframe):
        """
        Applies formatting styles to DataFrame content in an Excel sheet.

        Args:
            wb (Workbook): OpenPyXL workbook object.
            sheet_name (str): Name of the Excel sheet.
            start_row (int): Row to start formatting.
            dataframe (pd.DataFrame): DataFrame containing data to style.

        Returns:
            None
        """
        # Load the specified sheet
        sheet = wb[sheet_name]

        # Define border style
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Bold font for column headers
        bold_font = Font(bold=True)

        # Set text alignment to wrap text
        wrap_alignment = Alignment(wrap_text=True)


        # Get the number of rows and columns in the DataFrame
        n_rows, n_cols = dataframe.shape
        
        # Apply styles to the column headers (first row of the DataFrame or summary)
        for col_idx in range(1, n_cols + 1):
            cell = sheet.cell(row=start_row, column=col_idx)
            cell.font = bold_font  # Bold the column header
            cell.border = thin_border  # Apply border to the column header
            cell.alignment = wrap_alignment  # Enable wrap text for the header
        if n_rows < 10:
            end_row = start_row + n_rows
        
        else: end_row = start_row + n_rows + 1

        # Apply borders only to the range where the DataFrame has actual data
        for row_idx in range(start_row + 1, end_row):
            for col_idx in range(1, n_cols + 1):
                # Get the actual value from the DataFrame for this cell
                cell_value = dataframe.iloc[row_idx - start_row - 1, col_idx - 1]
                
                # Only apply border if the cell has non-NaN, non-empty data
                if pd.notna(cell_value) and cell_value != "":
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border  # Apply border to each non-empty cell
                    cell.alignment = wrap_alignment  # Enable wrap text for the header
    
    def __del__(self):
        """
        Cleans up resources upon object destruction.
        """
        object_name = "TdmsFileManager object"
        print(f"{object_name} is destroyed.")

'''##############################################
                   Unit testing
##############################################'''
if __name__ == "__main__":
    output_file_path = "Data/Tesla-Model3/2020-tesla-model3-uncertainty-analysis.xlsx"
    tdms_directory = "/home/debashis/Documents/Data/AMTL-Test-Data"
    accuracy_parameter1, accuracy_parameter2 = 0.35, 0.09

    test_id_list_category_1st = [62005016, 62005017, 62005018, 62005019, 62005020]
    test_id_list_category_2nd = [62006032, 62006033, 62006034, 62006035, 62006036]
    test_id_list_category_3rd, test_id_list_category_4th, test_id_list_category_5th = ([] for i in range(3))
    tdms_file_manager = TdmsFileManager(output_file_path, tdms_directory, accuracy_parameter1, accuracy_parameter2)
    tdms_file_manager.set_test_id_list(test_id_list_category_1st, test_id_list_category_2nd, test_id_list_category_3rd, test_id_list_category_4th, test_id_list_category_5th)