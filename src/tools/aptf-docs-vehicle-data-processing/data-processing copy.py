"""
**********************************************************************************

data-processing.py
Created by: Debashis Das
Argonne National Laboratory
Transportation and Power Systems Division

**********************************************************************************
  
Description:
------------
The methods available from this class are the following:
- get_cycles_list(config): Method to get list of cycles from config file
- is_description_row(row): Method dentify description rows
- dict_to_df(dictionary): Method to create pandas dataframe from dictionary
********
"""

import pandas as pd
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def get_cycles_list(config):
    """
    Method to get list of cycles from config file
    """
    cycles_list = []
    
    for cycle_type in config["CycleTypes"]:
        for key, cycle_type_values in cycle_type.items():
            for element in cycle_type_values:
                cycles_list.append(element)
            
    return cycles_list

def is_description_row(row):
    """
     Define a function to identify description rows
    """
    return pd.isna(row['Test Time']) and pd.isna(row['Date'])
    
def dict_to_df(dictionary):
    """
    Method to create pandas dataframe from dictionary 
    """
    data_list = []
    for key, entries in dictionary.items():
        for entry in entries:
            # Find the maximum length of lists in the entry
            max_length = max(len(v) for v in entry.values())
            for i in range(max_length):
                data_entry = {}
                for k, v in entry.items():
                    data_entry[k] = v[i] if i < len(v) else None
                data_entry['Category'] = key
                data_list.append(data_entry)

    df = pd.DataFrame(data_list)
    # Reorder columns to place 'Category' at the start
    columns = ['Category'] + [col for col in df.columns if col != 'Category']
    df = df[columns]
    return df  
     
def merge_category_cells(df, start_row, sheet):
    """
    Function to merge cells in the 'Category' column
    """
    current_category = None
    start_merge = None
    
    for idx, row in df.iterrows():
        category = row['Category']
        row_num = start_row + idx
        
        if category == current_category:
            continue
        
        if current_category is not None:
            end_merge = row_num
            sheet.merge_cells(start_row=start_merge, start_column=1, end_row=end_merge, end_column=1)
        current_category = category
        start_merge = row_num + 1
        
    # Merge the last group
    if start_merge is not None:
        sheet.merge_cells(start_row=start_merge, start_column=1, end_row=row_num + 1, end_column=1)


def update_dict(key, value, dictionary, dictionary_name):
        """
        Method to update msg_count_dictionary
        """
        if key in dictionary:
            dictionary[key].append(value)
        else:
            dictionary[key] = [value]
            
        # print(f"Data in {dictionary_name} is:\n{dictionary}")

      
def populate_dictionary(config, df, index_list, desired_subcycle_name, sub_phase_number, iteration_key, dictionary, dictionary_name):
    """
    Method to create lists dynamically and append requied data into those lists
    """
    
    # create lists dynamically
    lists_dict = {f"{field}": [] for field in config['DataFields']}
    
    for index in index_list:
        if df.loc[index, 'Cycle'] == desired_subcycle_name:
            for field in config['DataFields']:
                if field in df.columns:
                    lists_dict[f"{field}"].append(df.loc[index, field])
                # else:
                #     print(f"Field '{field}' does not exist in the dataframe. Skipping this field.")
                
            lists_dict["Sub-Phase"].append(sub_phase_number)             
    
    update_dict(iteration_key, lists_dict, dictionary, dictionary_name)           
         
def matching_cycles(config, df, cycles_list):
    """
    Method to find index number that matches desired drive cycle (e.g., MCT)
    """
    
    # create dictionaries dynamically
    cycle_type_dicts = {}

    for cycle_type in config["CycleTypes"]:
        for cycle_type_key, cycle_type_value in cycle_type.items():
            dict_name = f"{cycle_type_key}_dictionary" 
            cycle_type_dicts[dict_name] = {}                

    sub_cycle_names = config["SubCycleNames"]
    iteration_keys = config["IterationKeys"]
    max_sub_phase_number = config["MaxSubPhase"]
    sub_phase_number = 1
    sub_cycle_name_counter = 0
    iteration_index = 0
    desired_index_list  =[]
    sub_cycle_finding_status = True
        
    for index, row in df.loc[:].iterrows():
       
        if row['Cycle'] == sub_cycle_names[sub_cycle_name_counter]:
            sub_cycle_name_counter += 1
            sub_cycle_finding_status = True            
               
            if row['Cycle'] in cycles_list:
                desired_index_list.append(index)            
                
        else: sub_cycle_finding_status = False
                    
        if sub_cycle_finding_status and sub_cycle_name_counter == len(sub_cycle_names):
            sub_cycle_finding_status = False

            # iterating both cycle_type dictionary and cycle types simultaneouly
            for (dict_name, dictionary), cycle_type in zip(cycle_type_dicts.items(), config["CycleTypes"]):
                for key, values in cycle_type.items():
                    for sub_cycle_name in values:
                        if sub_phase_number > max_sub_phase_number:
                            sub_phase_number = 1
                        
                        if iteration_index < len(iteration_keys):
                            populate_dictionary(config, df, desired_index_list, sub_cycle_name, sub_phase_number, iteration_keys[iteration_index], dictionary, dict_name)
                        
                        else:
                            print("Not enough iteration keys for all cycle types")
                            break
                        
                        sub_phase_number += 1
                        
            iteration_index += 1
            
        if not(sub_cycle_finding_status)  and sub_cycle_name_counter > 1:
            sub_cycle_name_counter = 0
            desired_index_list.clear()
    
    return cycle_type_dicts
    

def main():
    
    # Read the config file into a json object:
    configFile = open("configuration.json", 'r')
    config = (json.load(configFile))
    configFile.close()
    
    output_file_path = config["OutputFileName"]
    output_sheet_name = config['OutputSheetName']
    
    data_frame = pd.read_excel(config["InputFileName"], sheet_name = config['InputSheetName'],  skiprows = config['NoOfSkipRows'])

    # Apply the filter to exclude description rows
    filtered_df = data_frame[~data_frame.apply(is_description_row, axis=1)]
    
    # # Reading column names
    # column_names = filtered_df.columns

    # # Display column names
    # print("Column names:", column_names)
    
    cycles_list = get_cycles_list(config)
    cycle_type_dicts = matching_cycles(config, filtered_df, cycles_list) 
 
    # for dictionary_name, dictionary in cycle_type_dicts.items():
    #     print(f"Data in {dictionary_name} is:\n{dictionary}\n\n")
       
    # Create a Pandas Excel writer using openpyxl as the engine
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        start_row = 0
        for dict_name, dictionary in cycle_type_dicts.items():
            # Convert dictionary to DataFrame
            df = dict_to_df(dictionary)
            
            # Add dictionary name before DataFrame
            worksheet = writer.book.create_sheet(output_sheet_name) if start_row == 0 else writer.sheets[output_sheet_name]
            worksheet.append([dict_name.split('_')[0].upper()]) if start_row == 0 else worksheet.cell(row=start_row + 1, column=1, value=dict_name.split('_')[0].upper())
            
            # Write the DataFrame to the sheet starting at the next row
            df.to_excel(writer, sheet_name=output_sheet_name, startrow=start_row + 1, index=False)
            
            # Update start_row for the next dictionary
            start_row += len(df) + 3

    # Load the workbook and access the sheet
    workbook = load_workbook(output_file_path)
    sheet = workbook[output_sheet_name]
    
    # Merge cells in 'Category' column for each DataFrame
    start_row = 2
    for dict_name, dictionary in cycle_type_dicts.items():
        df = dict_to_df(dictionary)
        merge_category_cells(df, start_row, sheet)
        start_row += len(df) + 3

    # Save the updated workbook
    workbook.save(output_file_path)
    print(f"File created successfully: {output_file_path}")

if __name__ == "__main__":
    main()  