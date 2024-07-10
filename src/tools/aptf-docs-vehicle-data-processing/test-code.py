
#Code to create two tables in two excel sheets

# import pandas as pd

# # Create the first DataFrame (Table 1)
# data1 = {
#     'Column1': [1, 2, 3],
#     'Column2': [4, 5, 6],
#     'Column3': [7, 8, 9]
# }
# df1 = pd.DataFrame(data1)

# # Create the second DataFrame (Table 2)
# data2 = {
#     'A': ['foo', 'bar', 'baz'],
#     'B': [10, 20, 30],
#     'C': [100, 200, 300]
# }
# df2 = pd.DataFrame(data2)

# # Define the path to save the new file
# output_file_path = 'two_tables.xlsx'

# # Create a Pandas Excel writer using XlsxWriter as the engine
# with pd.ExcelWriter(output_file_path, engine='xlsxwriter') as writer:
#     # Write each DataFrame to a specific sheet
#     df1.to_excel(writer, sheet_name='Table1', index=False)
#     df2.to_excel(writer, sheet_name='Table2', index=False)

# print(f"File created successfully: {output_file_path}")


######### Code to create two tables in one excel sheet

# import pandas as pd

# # Create the first DataFrame (Table 1)
# data1 = {
#     'Column1': [1, 2, 3],
#     'Column2': [4, 5, 6],
#     'Column3': [7, 8, 9]
# }
# df1 = pd.DataFrame(data1)

# # Create the second DataFrame (Table 2)
# data2 = {
#     'A': ['foo', 'bar', 'baz'],
#     'B': [10, 20, 30],
#     'C': [100, 200, 300]
# }
# df2 = pd.DataFrame(data2)

# # Define the path to save the new file
# output_file_path = 'two_tables_one_sheet.xlsx'

# # Create a Pandas Excel writer using openpyxl as the engine
# with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
#     # Write the first DataFrame to the sheet starting at cell A1
#     df1.to_excel(writer, sheet_name='Sheet1', startrow=0, index=False)
    
#     # Write the second DataFrame to the sheet starting at cell A6 (below the first table)
#     df2.to_excel(writer, sheet_name='Sheet1', startrow=len(df1) + 2, index=False)

# print(f"File created successfully: {output_file_path}")


##### Code to create variables dynamically

# Initialize an empty dictionary to hold the lists
# lists_dict = {}

# # Create and initialize the lists dynamically
# for i in range(1, 6):
#     list_name = f"list{i}"
#     lists_dict[list_name] = []

# # Append data to the lists dynamically
# for i in range(1, 6):
#     list_name = f"list{i}"
#     lists_dict[list_name].append(i * 10)  # Example data to append

# # Print the lists
# for list_name, data in lists_dict.items():
#     print(f"{list_name}: {data}")


# Method to create tables based on dictionary and save into xlsx file

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Define the nested dictionaries
udds_dictionary = {
    "1st": [
        {
            "Test ID [#]": [62009019],
            "Sub-Phase": [1],
            "Start SOC": [97.32648],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [9]
        },
        {
            "Test ID [#]": [62009019],
            "Sub-Phase": [2],
            "Start SOC": [92.908999],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [3356]
        },
        {
            "Test ID [#]": [62009021],
            "Sub-Phase": [1],
            "Start SOC": [16.392744],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [54955]
        },
        {
            "Test ID [#]": [62009021],
            "Sub-Phase": [2],
            "Start SOC": [11.244357],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [58159]
        }
    ],
    "2nd": [
        {
            "Test ID [#]": [62009024],
            "Sub-Phase": [2],
            "Start SOC": [97.382938],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [-1]
        },
        {
            "Test ID [#]": [62009024],
            "Sub-Phase": [1],
            "Start SOC": [92.884585],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [3389]
        },
        {
            "Test ID [#]": [62009026],
            "Sub-Phase": [2],
            "Start SOC": [16.778796],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [54810]
        },
        {
            "Test ID [#]": [62009026],
            "Sub-Phase": [1],
            "Start SOC": [16.778796],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [58002]
        }
    ],
    "3rd": [
        {
            "Test ID [#]": [62009059],
            "Sub-Phase": [1],
            "Start SOC": [97.212037],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [5]
        },
        {
            "Test ID [#]": [62009059],
            "Sub-Phase": [2],
            "Start SOC": [86.567359],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [7769]
        },
        {
            "Test ID [#]": [62009061],
            "Sub-Phase": [1],
            "Start SOC": [37.795017],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [40631]
        },
        {
            "Test ID [#]": [62009061],
            "Sub-Phase": [2],
            "Start SOC": [28.041464],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [46999]
        }
    ]
}

hwy_dictionary = {
    "1st": [
        {
            "Test ID [#]": [62009019],
            "Sub-Phase": [1],
            "Start SOC": [95.676982],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [1285]
        },
        {
            "Test ID [#]": [62009021],
            "Sub-Phase": [2],
            "Start SOC": [14.482317],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [56147]
        }
    ],
    "2nd": [
        {
            "Test ID [#]": [62009024],
            "Sub-Phase": [1],
            "Start SOC": [95.651042],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [1319]
        },
        {
            "Test ID [#]": [62009026],
            "Sub-Phase": [2],
            "Start SOC": [16.778796],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [55995]
        }
    ],
    "3rd": [
        {
            "Test ID [#]": [62009059],
            "Sub-Phase": [1],
            "Start SOC": [91.766752],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [4019]
        },
        {
            "Test ID [#]": [62009061],
            "Sub-Phase": [2],
            "Start SOC": [33.340915],
            "HV Batt (500A)\nStart \nWP1 [Wh]\n": [43586]
        }
    ]
}

def dict_to_df(dictionary):
    data_list = []
    for key, entries in dictionary.items():
        for entry in entries:
            data_entry = {k: v[0] for k, v in entry.items()}
            data_entry['Category'] = key
            data_list.append(data_entry)
    df = pd.DataFrame(data_list)
    # Reorder columns to place 'Category' at the start
    columns = ['Category'] + [col for col in df.columns if col != 'Category']
    df = df[columns]
    return df

# Convert dictionaries to DataFrames
udds_df = dict_to_df(udds_dictionary)
hwy_df = dict_to_df(hwy_dictionary)

# Define the path to save the new file
output_file_path = 'test_data_with_titles.xlsx'

# Create a Pandas Excel writer using openpyxl as the engine
with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
    # Add dictionary name before UDDS DataFrame
    worksheet = writer.book.create_sheet('Sheet1')
    worksheet.append(['UDDS'])
    
    # Write the UDDS DataFrame to the sheet starting at cell A2
    udds_df.to_excel(writer, sheet_name='Sheet1', startrow=1, index=False)
    
    # Add dictionary name before HWY DataFrame
    hwy_start_row = len(udds_df) + 3
    worksheet.cell(row=hwy_start_row, column=1, value='HWY')
    
    # Write the HWY DataFrame to the sheet starting below the HWY title
    hwy_df.to_excel(writer, sheet_name='Sheet1', startrow=hwy_start_row + 1, index=False)

# Load the workbook and access the sheet
workbook = load_workbook(output_file_path)
sheet = workbook['Sheet1']

# Function to merge cells in the 'Category' column
def merge_category_cells(df, start_row):
    current_category = None
    start_merge = None
    for idx, row in df.iterrows():
        category = row['Category']
        row_num = start_row + idx
        if category == current_category:
            continue
        if current_category is not None:
            end_merge = row_num
            sheet.merge_cells(start_row=start_merge, start_column=1, end_row=end_merge, end_column=1)
        current_category = category
        start_merge = row_num + 1
    # Merge the last group
    if start_merge is not None:
        sheet.merge_cells(start_row=start_merge, start_column=1, end_row=row_num + 1, end_column=1)

# Merge cells in 'Category' column for UDDS DataFrame
merge_category_cells(udds_df, 2)

# Merge cells in 'Category' column for HWY DataFrame
merge_category_cells(hwy_df, hwy_start_row + 2)

# Save the updated workbook
workbook.save(output_file_path)

print(f"File created successfully: {output_file_path}")
